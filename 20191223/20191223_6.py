# -*- coding: utf-8 -*-
# @Time : 2019/12/25 16:03
# @Author : 52595
# @File : 20191223_6.py
# @Python Version : 3.7.4
# @Software: PyCharm

from datetime import datetime
from urllib.parse import parse_qsl, urlparse

import requests
import json
import urllib3
import openpyxl
import prettytable

'''
--功能描述：
            提取12036查询页面中的车次信息，不用人为的录入数据。
--注意事项：
            1、需要获取页面的查询路径，即url_query字段的值。
            2、浏览器的cookies的值。
            3、使用狐火浏览器方便查看json格式中的内容。

--学习内容：
            1、根据url请求，获取页面中信息。
            2、json格式的转换以及json内容的读取
            备注：关键看怎么读取内容，因为如果换一个静态页面，自己是否也能进行爬取
            3、excel文件内容的读写操作

update history:
    20191225,打开excel文件后，执行程序包，由于软件已打开，导致程序出错。改为文件名自动生成
                知识点：（1）解析url,获取传入的参数值
                        （2）日期类使用，获取当前时间
                        （3）数据结构：列表、字典的使用
                        
                缺点：执行效率低。
'''


# 往Excel中写内容，参数1：excel写入的路径；参数2：excel工作表名称；参数3：需要写入的具体值
# 注意事项：如果执行代码前打开该Excel，则执行代码会报错
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


# 读取Excel中内容，参数1：读取excel中的路径；参数2：
def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


def get_station_name_code():
    # 获取当前站点代码对应的站点中文名  begin
    url = 'https://kyfw.12306.cn/otn/resources/js/framework/station_name.js?station_version=1.9130'
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    url_tmp = str(requests.get(url, verify=False).content, encoding='utf-8').replace("var station_names ='@",
                                                                                     '').replace(
        "';", '')
    url_tmp = list(url_tmp.split('@'))
    dct = locals()
    for i in url_tmp:
        dct[i.split('|')[2]] = i.split('|')[1]
    return dct
    # 获取当前站点代码对应的站点中文名  end


# 根据页面中信息，解析成json格式，然后返回车次信息
def show_ticket(url_query_value, cookies_value):
    # 获取完成，存放在字典中
    # 本地读取内容
    header = {
        'Cookie': cookies_value,
        'Host': 'kyfw.12306.cn',
        "Connection": "keep-alive",
        "X-Requested-With": "XMLHttpRequest",
        'Referer': 'https://kyfw.12306.cn/otn/leftTicket/init?linktypeid=dc'
        # Firefox表示火狐
        # 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:59.0) Gecko/20100101 Firefox/59.0'
        # Chrome表示谷歌
        # 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'
    }
    response = requests.get(url_query, headers=header)
    html = json.loads(response.content)
    # print(html)
    table_top = [" 车次 ", "出发车站", "到达车站", "出发时间", "到达时间", " 历时 "]
    # list中可以嵌套list
    tickets_excel = list()
    tickets_excel.append(table_top)
    name = [
        "station_train_code",
        "from_station_name",
        "to_station_name",
        "start_time",
        "arrive_time",
        "lishi"
    ]
    data = {
        "station_train_code": '',
        "from_station_name": '',
        "to_station_name": '',
        "start_time": '',
        "arrive_time": '',
        "lishi": ''
    }

    table = prettytable.PrettyTable([" 车次 ", "出发车站", "到达车站", "出发时间", "到达时间", " 历时 "])
    for i in html['data']['result']:
        # 将各项信息提取并赋值
        item = i.split('|')  # 使用“|”进行分割
        # split函数能将字符串根据某字符进行分割，然后返回一个数组
        data["station_train_code"] = item[3]  # 获取车次信息，在3号位置
        data["from_station_name"] = item[6]  # 始发站信息在6号位置
        data["to_station_name"] = item[7]  # 终点站信息在7号位置
        data["start_time"] = item[8]  # 出发时间在8号位置
        data["arrive_time"] = item[9]  # 抵达时间在9号位置
        data["lishi"] = item[10]  # 经历时间在10号位置
        # color = Colored()
        # data["note_num"] = color.white(item[1])
        # 如果没有信息，那么就用“-”代替
        for pos in name:
            if data[pos] == "":
                data[pos] = "-"
        tickets = []
        # tickets.append(tableTop)
        cont = list()
        cont.append(data)
        dic_name_code = get_station_name_code()
        # 循环的目的，数据处理，对于车次信息，换算成具体的车站名称  lwc,20191223
        for x in cont:
            tmp = []
            for y in name:
                # 起点站
                if y == "from_station_name":
                    # s = color.green(chezhan_names[data["from_station_name"]])
                    s1 = dic_name_code[data["from_station_name"]]
                    tmp.append(s1)
                # 终点站
                elif y == "to_station_name":
                    # s = color.red(chezhan_names[data["to_station_name"]])
                    s1 = dic_name_code[data["to_station_name"]]
                    tmp.append(s1)
                # 起始时间
                elif y == "start_time":
                    # s = color.green(data["start_time"])
                    s1 = data["start_time"]
                    tmp.append(s1)
                # 到达时间
                elif y == "arrive_time":
                    # s = color.red(data["arrive_time"])
                    s1 = data["arrive_time"]
                    tmp.append(s1)
                # 车次信息
                elif y == "station_train_code":
                    # s = color.yellow(data["station_train_code"])
                    s1 = data["station_train_code"]
                    tmp.append(s1)
                elif y == "lishi":
                    s1 = data["lishi"]
                    tmp.append(s1)
            # print(tmp)
        tickets_excel.append(tmp)
    return tickets_excel


# 根据url，获取传入参数，本程序中主要获取车次信息
def get_code_from_url(url_query_value):
    str_qsl = parse_qsl(urlparse(url_query).query)
    # 定义一个空的字典
    dict_train_code = dict()
    for str_qsl_temp in str_qsl:
        # 始发站
        if str_qsl_temp[0] == "leftTicketDTO.from_station":
            from_station_code = str_qsl_temp[1]
            dict_train_code['from_station_code'] = from_station_code
        # 终点站
        elif str_qsl_temp[0] == "leftTicketDTO.to_station":
            to_station_code = str_qsl_temp[1]
            dict_train_code['to_station_code'] = to_station_code
    return dict_train_code


# -------------------------具体的传参以及代码执行过程-------------------------------------------
# 车次信息输出Excel的路径
# 每次由于指定输出文件，由于打开，导致执行程序报错，因此根据时间作为文件名称  lwc,20191225
# 生成文件名
str_now_time = datetime.now().__format__('_%Y%m%d_%H%M_%S')
# 文件名
xmls_name = 'train_info' + str_now_time + '.xlsx'
# 存放路径
xmls_path = 'D:\\developmentSoft\\pythonWorkspace\\trainInfo\\'
# 拼接后的Excle路径以及名称
train_name_xlsx = xmls_path + xmls_name

# 注意更换此处内容
url_query = 'https://kyfw.12306.cn/otn/leftTicket/queryO?leftTicketDTO.train_date=2020-03-11&leftTicketDTO.from_station=SHH&leftTicketDTO.to_station=SZQ&purpose_codes=ADULT'
cookies = 'JSESSIONID=56528D1AD7E7FC3F275582FB65A0A187; RAIL_EXPIRATION=1582830052593; RAIL_DEVICEID=rhC5FZyKo4J2BaH1C7VftKza15SGap2FoZGvYN28oUU6lwpLcuaWlqQEew19BXMjwrDGbVe3kRvpiksLAlIyLYB2WbaP6oUrETbSiCgwgxkpNsUVMkDck4tHCoq8eGQE8_mN6Cegyjhsr5j1cl6Xol4bU4hybYAF; _jc_save_fromStation=%u4E0A%u6D77%2CSHH; _jc_save_toStation=%u6DF1%u5733%2CSZQ; _jc_save_fromDate=2020-03-11; _jc_save_toDate=2020-02-24; _jc_save_wfdc_flag=dc; BIGipServerotn=99615242.64545.0000; BIGipServerpool_passport=216859146.50215.0000; route=6f50b51faa11b987e576cdb301e545c4'
# 获取车站代码与车站名称
dic_station_code_and_name = get_station_name_code().copy()
# 获取url中的车站代码
dic_station_code = get_code_from_url(url_query).copy()

# 工作表名称
# sheet_name_xlsx = 'tab_1'
sheet_name_xlsx = dic_station_code_and_name[dic_station_code['from_station_code']] + "_" + dic_station_code_and_name[
    dic_station_code['to_station_code']]
# print(sheet_name_xlsx)

# 开始调用函数
# 1、获取页面中的数据，并提取车次信息并返回
value3 = show_ticket(url_query, cookies)

# 2、将获取到的数据写入Excel中，若需要读取Excel中的内容，则放开第3步的代码
# 注意事项，打开excel后，允许此程序则会报错 ，写入 20191224
# 文件名称自动生成   20191225
write_excel_xlsx(train_name_xlsx, sheet_name_xlsx, value3)

# 3、读取Excel
# read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)
