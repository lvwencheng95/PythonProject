# -*- coding: utf-8 -*-
# @Time : 2020/3/17 19:16
# @Author : 52595
# @File : nameTransform.py
# @Python Version : 3.7.4
# @Software: PyCharm
"""
功能说明：将中文名称转换为英文名称，示例：张三   转换后   zhangsan
问题1：写excle时，覆盖了之前内容，所以只能每次新建
"""
import datetime

import openpyxl


def read_excel_xlsx_return_list(path, sheet_name):
    xlsx_list = list()
    # 加载excel文档
    workbook = openpyxl.load_workbook(path)
    # 创建一个excel文档
    # workbook = openpyxl.workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    # 对行进行循环
    for row in sheet.rows:
        # 对某行的列进行循环
        for cell in row:
            xlsx_list.append(cell.value)
    return xlsx_list


def read_excel_xlsx_return_dict(path, sheet_name):
    xlsx_list = list()
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    xlsx_dict = dict()
    # 对行进行循环
    list_data = list()
    for row in sheet.rows:
        list_data.clear()
        # 对某行的列进行循环
        for cell in row:
            list_data.append(cell.value)
        xlsx_dict[list_data[0]] = list_data[1]
    return xlsx_dict


# 往Excel中写内容，参数1：excel写入的路径；参数2：excel工作表名称；参数3：需要写入的具体值
# 注意事项：如果执行代码前打开该Excel，则执行代码会报错
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    # 使用Workbook()方法会进行新建一个，然后保存会覆盖之前的
    # workbook = openpyxl.Workbook()
    # 使用load_workbook打开已有的
    workbook = openpyxl.load_workbook(path)
    # 默认会获取第一个sheet
    # sheet = workbook.active
    # 创建一个sheet，并进行命名
    sheet = workbook.create_sheet()
    sheet.title = sheet_name
    i = 1
    for value_key in value:
        sheet.cell(row=i + 1, column=2, value=value_key)
        sheet.cell(row=i + 1, column=3, value=value[value_key])
        i = i + 1
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


"""
    # 写入内容时，另外一种方式
    # 行
    for i in range(0, index):
        for j in range(0, 1):
            for value_key in value:
                # print(value_key)
                # print(value[value_key])
                sheet.cell(row=i + 1, column=j + 1, value=value_key)
                sheet.cell(row=i + 1, column=j + 2, value=value[value_key])
                value.pop(value_key)
                break
"""

# 获取单个汉字对应的拼音
dic = read_excel_xlsx_return_dict('E:\\pythonTestFile\\dataDictionary.xlsx', 'dictionary')
# print(dic)
dic_name = dict()
# 获取中中文名称
# 若需要读取则在此处更改文件的位置
list_name = read_excel_xlsx_return_list('E:\\pythonTestFile\\888.xlsx', 'name')
# 测试
# list_name = ['马云颍', '吴佳豪']
# print(list_name)
for name in list_name:
    temp_name = ''
    flag = '0'
    for temp in name:
        if temp in dic:
            temp_name = temp_name + dic[temp]
        else:
            temp_name = '0' + temp;
            break
    dic_name[name] = temp_name
# 生成文件名
str_now_time = datetime.datetime.now().__format__('_%Y%m%d_%H%M_%S')
# 文件名
xmls_name = 'name_info' + str_now_time + '.xlsx'
# 存放路径
xmls_path = 'D:\\developmentSoft\\pythonWorkspace\\nameInfo\\'
# 拼接后的Excle路径以及名称
name_xlsx = xmls_path + xmls_name
print('文件存放路径：  ' + name_xlsx)
name_xlsx = 'E:\\pythonTestFile\\12345678.xlsx'
write_excel_xlsx(name_xlsx, 'name_info', dic_name)
# print(dic_name)
'''
str = '吕文铖'
for temp in str:
    print(temp)
'''
