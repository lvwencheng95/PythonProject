# -*- coding: utf-8 -*-
# @Time : 2020/3/21 22:53
# @Author : 52595
# @File : Hanzi2Pinyin.py
# @Python Version : 3.7.4
# @Software: PyCharm
import binascii
import json
import re

import openpyxl

"""
学习内容;   1、编码格式的转换
           2、字符串的截取  a[x:y]
           3、字符串、列表、字典间的转换
           参考资料：https://www.cnblogs.com/who-care/p/9306800.html
"""
import requests


# 读取excel中内容，使用list结构返回
def read_excel_xlsx_return_list(path, sheet_name):
    xlsx_list = list()
    # 加载excel文档
    workbook = openpyxl.load_workbook(path)
    # 创建一个excel文档
    # workbook = openpyxl.workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    # 对行进行循环
    for row in sheet.rows:
        # 对某行的列进行循环
        for cell in row:
            if cell.value is None:
                continue
            xlsx_list.append(cell.value)
    return xlsx_list


# 注意事项：如果执行代码前打开该Excel，则执行代码会报错
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    # 使用Workbook()方法会进行新建一个，然后保存会覆盖之前的
    # workbook = openpyxl.Workbook()
    # 使用load_workbook打开已有的
    workbook = openpyxl.load_workbook(path)
    # 默认会获取第一个sheet
    # sheet = workbook.active
    # 创建一个sheet，并进行命名
    sheet = workbook.create_sheet()
    sheet.title = sheet_name
    i = 1
    for value_key in value:
        sheet.cell(row=i + 1, column=2, value=value_key)
        sheet.cell(row=i + 1, column=3, value=value[value_key])
        i = i + 1
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


url_query = 'https://bihua.51240.com/web_system/51240_com_www/system/file/bihua/get_0/?shi_fou_zi_dong=1&cache_sjs1=20031901'
cooky11 = 'Hm_lvt_fbe0e02a7ffde424814bef2f6c9d36eb=1584433904,1584520595,1584541953,1584799808; __gads=ID=a14a729f79924a24:T=1584520596:S=ALNI_MZGE-F7i75SaXjljA_Kzenbzmxn-w; Hm_lpvt_fbe0e02a7ffde424814bef2f6c9d36eb=1584800309; c_y_g_j=bihua%2Czhongwenzhuanpinyin'
header = {
    'Cookie': cooky11,
    'Host': 'bihua.51240.com',
    'Referer': 'https://bihua.51240.com/e5bca0__bihuachaxun/',
    # Firefox表示火狐
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:59.0) Gecko/20100101 Firefox/59.0'
    # Chrome表示谷歌
    # 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'
}
dict_hanzi = dict()
# 读取excel中的汉字，另一种方式，可读取txt或word中的内容
list_hanzi = read_excel_xlsx_return_list('E:\\pythonTestFile\\hanzi.xlsx', 'hanZi')
# list_hanzi = ['我', '是', '中', '国', '人']
for hanzi in list_hanzi:
    # 需使用decode进行解码，否则获取的编码前带上b
    hanzi_code = binascii.b2a_hex(hanzi.encode('utf-8')).decode('utf-8')
    params = {
        'font': hanzi_code
    }
    response = requests.get(url_query, headers=header, params=params)
    response_content = response.content.decode("utf-8")
    # 由于返回不是json格式内容，因此无法进行转换,因此进行截取
    """
    l_index = response_content.index('[', 0, len(response_content))
    r_index = response_content.index('}', 0, len(response_content))
    # 将字符串转换为列表list
    content = eval(response_content[l_index:r_index])
    """
    # print(response_content)
    pattern = re.compile(r'(?<=\[)[^}]*(?=\])')  # 查找[]中的内容,[^}]表示非}内容
    # 返回一个list
    result_list = pattern.findall(response_content)
    content = eval(result_list[0])
    index_list = len(content)
    dict_hanzi[hanzi] = content[index_list - 1]
# print(dict_hanzi)
write_excel_xlsx('E:\\pythonTestFile\\hanzi.xlsx', 'hanZi2PinYin', dict_hanzi)

